from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from controleCEQ.models import Abastecimento, Entrada
from ativos.models import Obras
from openpyxl import Workbook
from io import BytesIO
from django.db.models import Sum
from datetime import datetime
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import json
from django.contrib import messages
from django.contrib.messages import constants

# Create your views here.



def obra(request,id):  #Recebi o id da página html lista_obra: href="/obra/{{obra.id}}"
    
    obra = Obras.objects.get(id=id)#objeto - obra do usuario logado
    obra_name = obra
    saidas = Abastecimento.objects.filter(obra=obra_name)
    entradas = Entrada.objects.filter(obra=obra_name)
    obras = Obras.objects.filter(nome=obra_name)
    user = request.user



    return render(request, 'obra.html', {'obra': obra, 
                                         'saidas':saidas, 
                                         'entradas':entradas, 
                                         'obras':obras, 
                                         'user':user, 
                                         'obra_user':obra})

def status(request, id):
    # obra = Obra.objects.get(id=id)
    abastecimento = Abastecimento.objects.get(id=id)
    obra = Obras.objects.get(nome=abastecimento.obra)
    abastecimento.status = not abastecimento.status #toda vez que clicka, inverto o atual
    abastecimento.save()
    id_user = request.user.id
    if str(request.user.status) == 'c':
        return redirect("/ceq/painel_obras")
    else:      
        return redirect(f"/obra/{obra.id}")

def comentario(request, id):
    abastecimento = Abastecimento.objects.get(id=id)
    
    obra = Obras.objects.get(nome=abastecimento.obra)
    comment = request.POST.get('comentario')

    abastecimento.observacao = comment

    abastecimento.save() 
    

    id_user = request.user.id

    if request.user.status == 'c':
        return redirect('/ceq/painel_obras')
    else:
        return redirect((f"/obra/{obra.id}"))


def exportexcel(request):
    id_user = request.user.id

    obra = Obras.objects.filter(usuario=id_user)#objeto - obra do usuario logado
    obra_name = obra[0] #Nome da obra para filtrar nas demais classes.
    saidas = Abastecimento.objects.filter(obra=obra_name)
    entradas = Entrada.objects.filter(obra=obra_name)
    obras = Obras.objects.filter(nome=obra_name)
    sd = []
    sds = [['ID', 'DATA', 'OBRA', 'EQUIPAMENTO', 'LITROS', 'HORIMETRO', 'OPERADOR']]
    en = []
    ens = [['ID', 'DATA ENTREGA', 'QUANTIDADE', 'NF', 'DATA EMISSÃO', 'VALOR LITRO', 'VALOR TOTAL']] 

    for entrada in entradas:
        en.append(entrada.id)
        en.append(entrada.data_entrega)
        en.append(entrada.quantidade)
        en.append(entrada.nota_fiscal)
        en.append(entrada.data_nf)
        en.append(entrada.preco_unitario)
        en.append(entrada.preco_total)
        ens.append(en)
        en = []


    for saida in saidas:
        sd.append(saida.id)
        sd.append(saida.data)
        sd.append(saida.obra.nome)
        sd.append(saida.equipamento.prefixo)
        sd.append(saida.litros)
        sd.append(saida.horimetro)
        sd.append(saida.operador)
        sds.append(sd)
        sd = []
    print(sds)
    print(f'okok{sds}')

    todays_date = datetime.today().date()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f"attachment; filename=diesel {obra_name.nome} {todays_date}.xlsx"

    wb = Workbook()
    ws = wb.active

        ############### style border ##########################
    border = Border(left=Side(border_style='thin',
                              color='FF000000'),
                    top=Side(border_style='thin',
                              color='FF000000'),
                    bottom=Side(border_style='thin',
                              color='FF000000'),
                    right=Side(border_style='thin',
                              color='FF000000'))
        ############### style borde r##########################



    title_saldo_cell = ws.cell(row=2, column=1, value="SALDO ATUAL")
    saldo_valor_cell = ws.cell(row=3, column=1, value=obra_name.saldo)
    
    start_row = 6

    title_entrada = ws.cell(row=start_row-1, column=1, value="ENTRADAS")
    for row_index, list in enumerate(ens, start = start_row):
        for col_index, value in enumerate(list, start = 1):
            ws.cell(row = row_index, column = col_index, value=value).border = border

    title_e_total = ws.cell(row=start_row+len(ens), column=2, value="Total Entradas")
    value_e_total = ws.cell(row=start_row+len(ens), column=3, value=entradas.aggregate(Sum('quantidade'))['quantidade__sum'])


    title_saida = ws.cell(row=start_row+len(ens)+2, column=1, value='SAIDAS')
    for row_index, list in enumerate(sds, start=start_row + len(ens)+3):
        for col_index, value in enumerate(list, start =1):
            ws.cell(row = row_index, column = col_index, value=value).border =border

    title_s_total = ws.cell(row=start_row+len(ens)+len(sds)+3, column=4, value="Total saídas")
    value_s_total = ws.cell(row=start_row+len(ens)+len(sds)+3, column=5, value=saidas.aggregate(Sum('litros'))['litros__sum'])


    #######################STYLE PLANILHA###########s###############

    font_title = Font(name='Calibri',
                size=20,
                color='FF000000')
    
    font_label = Font(name='Calibri',
                      size=13,
                      bold=True)
    
    font_data = Font(name='Calibri',
                      size=11)
    fill_title = PatternFill("solid", fgColor="DDDDDD")
    fill_label = PatternFill("solid", fgColor="DDDDDD")


    obra_title = ws.cell(row=2, column=2, value=obra_name.nome)
    obra_title.font = Font(name='Calibri',size=24)
    obra_title.alignment = Alignment(horizontal='center', vertical='center')
    obra_title.border = border
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=7)

    title_saldo_cell.border = border
    title_saldo_cell.font = font_title
    saldo_valor_cell.border = border
    saldo_valor_cell.font = font_title

    title_entrada.font = font_title
    title_entrada.border = border
    ws.merge_cells(start_row=start_row-1, start_column=1, end_row = start_row-1, end_column=7)
    title_entrada.alignment = Alignment(horizontal='center')
    title_entrada.fill = PatternFill(start_color="00CCCCFF", end_color="00CCCCFF", fill_type="solid")

    title_e_total.font = font_label
    title_e_total.border = border
    value_e_total.font = font_label
    value_e_total.border = border
    
    title_saida.font = font_title
    title_saida.alignment = Alignment(horizontal='center')
    title_saida.border = border
    ws.merge_cells(start_row=start_row+len(ens)+2, start_column=1, end_row=start_row+len(ens)+2, end_column=7)
    title_saida.fill = PatternFill(start_color="00CCCCFF", end_color="00CCCCFF", fill_type="solid")

    title_s_total.font = font_label
    title_s_total.border = border
    value_s_total.font = font_label
    value_s_total.border = border

    for column in range(1,8):
        x = ws.cell(row=start_row, column=column)
        x.font = font_label
        y = ws.cell(row=start_row+len(ens)+3, column=column)
        y.font = font_label
    


 

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20   
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20  
    ws.column_dimensions['E'].width = 15 
    ws.column_dimensions['F'].width = 15          
    ws.column_dimensions['G'].width = 20 

    wb.save(response)

    #######################STYLE PLANILHA##########################
    messages.add_message(request, constants.SUCCESS, "Arquivo baixado com sucesso!" )
    return response

def testegrafico(request):

    valor = 5000

    return JsonResponse({'valor':valor})

def renderiza_grafico(request):

    valor = 5000

    return JsonResponse({'valor':valor})

def lista_obras(request):
    obras = Obras.objects.filter(usuario=request.user)
    meses = [1,2,3,4,5,6,7,8,9,10,11,12]
    consumo_meses = []
    consumo_mes = []
    for obra in obras:
        for mes in meses:
            saida = Abastecimento.objects.filter(data__month=mes).filter(obra=obra).aggregate(Sum('litros'))['litros__sum']
            if saida == None:
                saida = 0
            else:pass
            consumo_mes.append(saida)
        consumo_meses.append(consumo_mes)
        consumo_mes = []


    
    zipped_segments = zip(obras, consumo_meses)
    print(consumo_meses)
    

    return render(request, 'lista_obras.html', {'obras':obras, 'consumo_meses':consumo_meses, 'zipped_segments':zipped_segments})

