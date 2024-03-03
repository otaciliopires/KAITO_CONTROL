from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from .models import Abastecimento, Tanque, Transferencia, Entrada, Saldo
from ativos.models import Equipamentos, Obras
from autenticacao.models import Usuario
import datetime
from datetime import date, datetime
from django.db.models import Sum, Avg, Max
from . utils import nonetest, grafico_vunit
import openpyxl
import json
from operator import itemgetter
from django.contrib import messages
from django.contrib.messages import constants



""" Para filtrar uma ForeignKey foi necessário adquirir o valor.id do respectivo item. O filtro da ForeignKey requisita que você
faça isso. No value do formulário foi pedido o equipamento.id e a obra.id  --  em seguida, foi utilizado o 
filter(carateristica_objeto__in=lista_variável -- Em que a lista variável é uma lista de id que quer filtrar da ForeignKey"""


@login_required(login_url='/auth/login/')
def home(request):
    
    if request.method == 'GET' and request.user.status=='c':

        
        user = request.user
        obra_user=Obras.objects.filter(usuario=user.id)
        tanques = Tanque.objects.all()
        obras = Obras.objects.all()
        equipamentos = Equipamentos.objects.all()

        # delete_all = Entrada.objects.all()
        # delete_all.delete()
        
        #Método para cadastrar os equipamentos

        # excel = "media/fotos/equipamentos2.xlsx"
        # workbooks = openpyxl.load_workbook(excel)
        # equipamentss = workbooks['Equipamentos']
        # total_list_x=[]
        # list_x = []
        # for i in equipamentss.iter_rows(min_row=2,values_only=True):
        #         equip = i[:4]
        #         print(equip)
        #         if equip[0] == None:
        #                 break
        #         else:
        #             total_list_x.append(equip) 
        # for i in total_list_x:
        #     print(i[3]) 
        #     cadastro_equipamentoss = Equipamentos(prefixo=i[0],
        #                                           descricao=i[0],
        #                                           tipo=i[3],
        #                                           proprietario=i[2],
        #                                           horímetro=0

        #     )

        #     cadastro_equipamentoss.save()
        # x =Equipamentos(prefixo='AB-01',
        #                 descricao='AUTOBETONEIRA',
        #                 tipo='T',
        #                 proprietario='CONTRUTORA ROCHA',
        #                 horímetro=0)

        # x.save()
                    
        # x = Equipamentos.objects.all()
        # x.delete()
        





        
        # Método para atualizar saldo da obra.
        for obra in obras:
            
            saldo = Entrada.objects.filter(obra=obra)  and Abastecimento.objects.filter(obra=obra)
            # if saldo == None:                
            #     obra.saldo = 0
            # else:
            #     obra.saldo = Entrada.objects.filter(obra=obra).aggregate(Sum('quantidade'))['quantidade__sum'] - Abastecimento.objects.filter(obra=obra).aggregate(Sum('litros'))['litros__sum']
            # obra.save()

        #Método para atualizar saldo dos tanques e lançar valores no frontend
        for tanque in tanques:
            if tanque.tipo == "F":
                tanque.estoque = nonetest(Entrada.objects.filter(tanque=tanque).aggregate(Sum('quantidade'))['quantidade__sum']) - nonetest(Abastecimento.objects.filter(tanque=tanque).aggregate(Sum('litros'))['litros__sum']) - nonetest(Transferencia.objects.filter(fixo=tanque).aggregate(Sum('litros'))['litros__sum'])
            if tanque.tipo == 'M':
                tanque.estoque = nonetest(Transferencia.objects.filter(movel=tanque).aggregate(Sum('litros'))['litros__sum']) - nonetest(Abastecimento.objects.filter(tanque=tanque).aggregate(Sum('litros'))['litros__sum'])
            
            tanque.save()
            if tanque.prefixo == 'TC-01':
                tc01 = [tanque.estoque, round(100*tanque.estoque/15000),1]
                
            elif tanque.prefixo == 'TC-02':
                tc02 = [tanque.estoque, round(100*tanque.estoque/30000, 1)]
            elif tanque.prefixo == 'CA-01.1':
                ca01 = [tanque.estoque, round(100*tanque.estoque/4200,1)]
            elif tanque.prefixo == 'CA-02.1':
                ca02 = [tanque.estoque, round(100*tanque.estoque/4500,2)]
        estoque_total = tc01[0]+tc02[0]+ca01[0]+ca02[0]
        porcent_estoque = round(100*(estoque_total/(15000+30000+4200+4500)),1)
        #Alterações horímetro equipamento
        for equipamento in equipamentos:

            equipamento.save()

        #criação de gráfico semanal no frontend
        ano = datetime.today().year
        mes = datetime.today().month
        meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']

        mes_atual = request.POST.get('mes')
        if mes_atual == None:
            mes_atual = meses[mes-1]
        mes = meses.index(mes_atual)+1
        print(mes_atual, mes)        

        data_inicio_1 = datetime(ano,mes,1)
        data_fim_1 = datetime(ano, mes, 8)
        data_inicio_2 = datetime(ano,mes,9)
        data_fim_2 = datetime(ano, mes, 16)
        data_inicio_3 = datetime(ano,mes,17)
        data_fim_3 = datetime(ano, mes, 24)
        data_inicio_4 = datetime(ano,mes,25)
        if mes == 'Janeiro' or mes =='Março' or mes =='Maio' or mes =='Julho' or mes =='Agosto' or mes =='Outubro' or mes =='Dezembro':
            data_fim_4 = datetime(ano, mes, 31)
        if mes == 'Abril' or mes =='Junho' or mes =='Setembro' or mes =='Novembro':
            data_fim_4 = datetime(ano, mes, 30)
        else:
            data_fim_4 = datetime(ano, mes, 28)



        entradas_1 = Entrada.objects.filter(data_entrega__range=[data_inicio_1, data_fim_1]).aggregate(Sum('quantidade'))['quantidade__sum']
        entradas_2 = Entrada.objects.filter(data_entrega__range=[data_inicio_2, data_fim_2]).aggregate(Sum('quantidade'))['quantidade__sum']
        entradas_3 = Entrada.objects.filter(data_entrega__range=[data_inicio_3, data_fim_3]).aggregate(Sum('quantidade'))['quantidade__sum']
        entradas_4 = Entrada.objects.filter(data_entrega__range=[data_inicio_4, data_fim_4]).aggregate(Sum('quantidade'))['quantidade__sum']
        entradas = json.dumps([entradas_1, entradas_2, entradas_3, entradas_4])

        saidas_1 = Abastecimento.objects.filter(data__range=[data_inicio_1, data_fim_1]).aggregate(Sum('litros'))['litros__sum']
        saidas_2 = Abastecimento.objects.filter(data__range=[data_inicio_2, data_fim_2]).aggregate(Sum('litros'))['litros__sum']
        saidas_3 = Abastecimento.objects.filter(data__range=[data_inicio_3, data_fim_3]).aggregate(Sum('litros'))['litros__sum']
        saidas_4 = Abastecimento.objects.filter(data__range=[data_inicio_4, data_fim_4]).aggregate(Sum('litros'))['litros__sum']
        saidas = json.dumps([saidas_1, saidas_2, saidas_3, saidas_4])   

        lista_obras = []
        lista_consumo = []

        
        #criação de gráfico obras no frontend
        for obra in obras:
            consumo_mensal_obras = Abastecimento.objects.filter(data__month=mes).filter(data__year=ano).filter(obra=obra).aggregate(Sum('litros'))['litros__sum']
            lista_obras.append(obra.nome)
            if consumo_mensal_obras == None:
                consumo_mensal_obras = 0
            lista_consumo.append(consumo_mensal_obras)
        
        list_obras=(lista_obras)
        list_consumo = json.dumps(lista_consumo)

        equipamentos = Equipamentos.objects.all()
        list_equip = []
        saidas_equip = []
        for equipamento in equipamentos:
            list_equip.append(equipamento.prefixo)
            saida_equip = Abastecimento.objects.filter(data__month=mes).filter(data__year=ano).filter(equipamento=equipamento).aggregate(Sum('litros'))['litros__sum']
            if saida_equip == None:
                saida_equip = 0
            saidas_equip.append(saida_equip)

            
        #LISTA EQUIPAMENTOS MAIOR CONSUMO
        dict_equip = dict(zip(list_equip,saidas_equip))
        sort_dict = dict(sorted(dict_equip.items(), key=itemgetter(1), reverse=True))
        sort_d = dict(list(sort_dict.items())[:5])


        return render(request, 'home.html', {'tanques':tanques, 
                                             'obras': obras, 
                                             'meses':meses,
                                             'equipamentos':equipamentos, 
                                             'user':user, 
                                            #  'obra_user':obra_user[0], 
                                             'tc01':tc01,
                                             'tc02':tc02,
                                             'ca01':ca01,
                                             'ca02':ca02,
                                             'estoque_total':estoque_total,
                                             'porcent_estoque':porcent_estoque,
                                             'mes_atual':mes_atual,
                                             'entradas':entradas,
                                             'saidas':saidas,
                                             'list_obras':list_obras,
                                             'list_consumo':list_consumo,
                                             'sort_dict':sort_d,})
    
    if request.method == 'POST':
        form_saidas = request.POST.get('form_saidas')
        form_transferencias = request.POST.get('form_transferencias')
        form_entradas = request.POST.get('form_entradas')
        form_test = request.POST.get('form_test')

# método acima é para quando for necessário selecionar um form específico em um html com mais de um form

        if form_saidas:

            tanque_id = request.POST.get('tanque_id')
            tanque = Tanque.objects.get(id=tanque_id)
            obra_id = request.POST.get('obra') #Método para buscar o id de uma ForeignKey
            obra = Obras.objects.get(id=obra_id)
            equipamento_id = request.POST.get('equipamento')
            equipamento =Equipamentos.objects.get(id=equipamento_id)
            
            contador_inicial = request.POST.get('contador_inicial')
            contador_final = request.POST.get('contador_final')
            litros = request.POST.get('saida_litros')
            horimetro = request.POST.get('horimetro')
            operador = request.POST.get('operador')
            data = date.today()
            num_saida = Abastecimento.objects.aggregate(Max('numero'))
            num_saida = (num_saida['numero__max'] + 1)



            #lançamento abastecimentos:
            abastecimento = Abastecimento(litros=litros,
                                          contador_inicio=contador_inicial,
                                          contador_fim=contador_final,
                                          horimetro=horimetro,
                                          data=data,
                                          tanque=tanque,
                                          obra=obra,
                                          equipamento=equipamento,
                                          operador=operador,
                                          colaborador= request.user,
                                          numero=num_saida)
            try:
                abastecimento.save()

                messages.add_message(request, constants.SUCCESS, "Abastecimento laçado com sucesso!" )
                return redirect("/ceq/home")
            except:
                messages.add_message(request, constants.ERROR, "ERRO AO LANÇAR O ABASTECIMENTO" )
                return redirect("/ceq/home")




            print(f"{tanque}, {obra}, {equipamento}, {contador_inicial}, {contador_final}, {type(litros)},{horimetro}, {operador}")
            return HttpResponse(f"{data}, {equipamento}, {contador_inicial}, {contador_final}, {litros},{horimetro}, {operador} -- ,tanque:{tanque}  tanque.saldo: {tanque.estoque} -- obra:{obra}, saldo:{obra.saldo}")
        
        
        if form_entradas:

            tanque_id = request.POST.get('tanque_id')
            tanque = Tanque.objects.get(id=tanque_id)
            obra_id = request.POST.get('obra_id')
            obra = Obras.objects.get(id=obra_id)
            data_emissao = request.POST.get('data_nf')
            data_entrega = request.POST.get('data_entrega')
            fornecedor = request.POST.get('fornecedor')
            nota_fiscal = request.POST.get('NF')
            valor_litro = request.POST.get('valor_litro')
            quantidade_litros = request.POST.get('quantidade')
            valor_total = float(valor_litro) * int(quantidade_litros)
            num_entrada = Entrada.objects.aggregate(Max('numero'))
            num_entrada = num_entrada['numero__max']+1

            #lançamento entradas:
            entrada = Entrada(tanque=tanque,
                              nota_fiscal=nota_fiscal,
                              fornecedor=fornecedor,
                              data_nf=data_emissao,
                              data_entrega=data_entrega,
                              obra=obra,
                              quantidade=quantidade_litros,
                              preco_unitario=valor_litro,
                              preco_total=valor_total,
                              colaborador=request.user,
                              numero=num_entrada
                              )
            entrada.save()

            #Método para atualizar saldo dos tanques.
            total_entradas = nonetest(Entrada.objects.filter(tanque=tanque).aggregate(Sum('quantidade'))['quantidade__sum'])
            total_saidas = nonetest(Abastecimento.objects.filter(tanque=tanque).aggregate(Sum('litros'))['litros__sum'])
            total_transferências = nonetest(Transferencia.objects.filter(fixo=tanque).aggregate(Sum('litros'))['litros__sum'])
            tanque.estoque = total_entradas - total_saidas - total_transferências
            tanque.save()
            
            #Método para atualizar saldo da obra.
            total_entradas = nonetest(Entrada.objects.filter(obra=obra).aggregate(Sum('quantidade'))['quantidade__sum'])
            total_saidas = nonetest(Abastecimento.objects.filter(obra=obra).aggregate(Sum('litros'))['litros__sum'])
            obra.saldo = total_entradas - total_saidas
            obra.save()
            print(f"obra:{obra} ---saldo_obra = {obra.saldo} --- estoque_tanque = {tanque.estoque}")
            return HttpResponse(f"{tanque}, {obra_id}, {type(data_emissao)}, {type(data_entrega)}, {fornecedor}, {nota_fiscal},{valor_litro}, {quantidade_litros}")

        
        if form_transferencias:
            tanque_fixo_id = request.POST.get('tanque_fixo_id')
            tanque_fixo = Tanque.objects.get(id=tanque_fixo_id)
            tanque_movel_id = request.POST.get('tanque_movel_id')
            tanque_movel = Tanque.objects.get(id=tanque_movel_id)

            contador_inicial = request.POST.get('contador_inicio')
            contador_final = request.POST.get('contador_fim')
            litros = float(contador_final) - float(contador_inicial)
            contador_comboio = request.POST.get('contador_comboio')
            data = date.today()

            #lançamento transferências:
            transferencia = Transferencia(fixo=tanque_fixo,
                                          movel=tanque_movel,
                                          contador_inicio=float(contador_inicial),
                                          contador_fim=float(contador_final),
                                          litros=litros,
                                          contador_comboio=float(contador_comboio),
                                          colaborador=request.user,
                                          data=date.today())
            
            transferencia.save()
            #Método para atualizar saldo dos tanques.
            tanque_fixo.estoque = nonetest(Entrada.objects.filter(tanque=tanque_fixo).aggregate(Sum('quantidade'))['quantidade__sum']) - nonetest(Abastecimento.objects.filter(tanque=tanque_fixo).aggregate(Sum('litros'))['litros__sum']) - nonetest(Transferencia.objects.filter(fixo=tanque_fixo).aggregate(Sum('litros'))['litros__sum'])
            tanque_movel.estoque = nonetest(Transferencia.objects.filter(movel=tanque_movel).aggregate(Sum('litros'))['litros__sum']) - nonetest(Abastecimento.objects.filter(tanque=tanque_movel).aggregate(Sum('litros'))['litros__sum'])
                
            
            tanque_fixo.contador = float(contador_final)
            tanque_movel.contador = float(contador_comboio)
            tanque_fixo.save()
            tanque_movel.save()

            print(f"{tanque_fixo.estoque}, {tanque_movel.estoque}")

            return HttpResponse(f"{tanque_fixo}, {tanque_movel}, {type(contador_inicial)}, {contador_final}, {contador_comboio}, LITROS:{litros}")

    else: return HttpResponse('<h1>Acesso negado</h1>')

@login_required(login_url='/auth/login/')
def saidas(request):
 if request.user.status=='c':
        
    obras = Obras.objects.all()
    equipamentos = Equipamentos.objects.all()
    list_equipamentos = []
    list_obras =[]

    for e in equipamentos:list_equipamentos.append(e)
    for o in obras: list_obras.append(o)
    

    data_inicio = request.POST.get('data_inicio')
    if data_inicio == None or data_inicio == '': datetime.strptime('2020-01-02', '%Y-%m-%d').date()
    else: data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        
    data_fim = request.POST.get('data_fim')
    if data_fim == None or data_fim == '': data_fim = date.today()
    else: data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()


    filtro_obras = request.POST.getlist('obra')
    filtro_equipamento = request.POST.getlist('equipamento')
    
    if data_inicio or data_fim or filtro_equipamento or filtro_obras:
        if not data_inicio:
            data_inicio = date(2020,1,1)
        if not data_fim:
            data_fim = date.today()
        if not filtro_equipamento:
            filtro_equipamento = list_equipamentos
        if not filtro_obras:
            filtro_obras= list_obras


        saidas = Abastecimento.objects.filter(data__range=[data_inicio, data_fim]).filter(equipamento__in=filtro_equipamento).filter(obra__in=filtro_obras).order_by('numero')
        total_saidas = Abastecimento.objects.filter(data__range=[data_inicio, data_fim]).filter(equipamento__in=filtro_equipamento).filter(obra__in=filtro_obras).aggregate(Sum('litros'))['litros__sum']

    else:
        saidas = Abastecimento.objects.all().order_by('numero')
    user = request.user
    obra_user=Obras.objects.filter(usuario=user.id)
    print(type(data_fim), data_fim)
    print(type(data_inicio), data_inicio)
    print(total_saidas)
    
    return render(request, 'saidas.html', {'saidas': saidas, 
                                           'obras': obras, 
                                           'equipamentos':equipamentos, 
                                           "user":user, 
                                           'total_saidas':total_saidas})
 
 elif request.user.status == 'o': 
    return HttpResponse('acesso negado')

@login_required(login_url='/auth/login/')
def entradas(request):
  if request.user.status=='c':

    obras = Obras.objects.all()

    list_obras = []
    for i in obras: list_obras.append(i)

    #filtro
    filtro_obra = request.POST.getlist('obra')
    data_inicio = request.POST.get('data_inicio')
    data_fim = request.POST.get('data_fim')

    if data_inicio or data_fim or filtro_obra:
        if not data_inicio:
            data_inicio = date(2020,1,1)
        elif data_inicio == None:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        if not data_fim:
            data_fim = date.today()
        elif data_fim == None:
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
        if not filtro_obra:
            filtro_obra = list_obras

    if request.method == "GET":
        if not data_inicio:
            data_inicio = date(2020,1,1)
        elif data_inicio == None:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        if not data_fim:
            data_fim = date.today()
        elif data_fim == None:
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
        if not filtro_obra:
            filtro_obra = list_obras

        
        entradas = Entrada.objects.filter(data_entrega__range=[data_inicio, data_fim], obra__in=filtro_obra).order_by('numero')
        total_entradas = Entrada.objects.filter(data_entrega__range=[data_inicio, data_fim], obra__in=filtro_obra).aggregate(Sum('quantidade'))['quantidade__sum']
    else:
        entradas = Entrada.objects.all().order_by('numero')
        total_entradas = Entrada.objects.filter(data_entrega__range=[data_inicio, data_fim], obra__in=filtro_obra).aggregate(Sum('quantidade'))['quantidade__sum']


    print(f"{filtro_obra} and {type(filtro_obra)}")
    print(f"{data_inicio} and {type(data_inicio)}")
    user = request.user
    obra_user=Obras.objects.filter(usuario=user.id)
    print(entradas)

    return render(request, 'entradas.html', {'entradas':entradas, 
                                             'obras': obras, 
                                             "user":user, 
                                             'total_entradas':total_entradas})
  else: return HttpResponse("<h1>Acesso negado</h1>")

def transferencias(request):
 if request.user.status == "c":

    transferencias = Transferencia.objects.all()
    tanque_fixo = Tanque.objects.filter(tipo='F')
    tanque_movel = Tanque.objects.filter(tipo='M')

    print(tanque_movel)
    data_inicio = request.POST.get('data_inicio')
    data_fim = request.POST.get('data_fim')
    tanque = request.POST.getlist('tanque_fixo')
    comboio = request.POST.getlist('tanque_movel')
    
    if data_inicio or data_fim or tanque or comboio:
        if not data_inicio:
            data_inicio = date(2022,1,1)
        else:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        if not data_fim:
            data_fim = date.today()
        else:
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
        if not tanque:
            tanque = tanque_fixo
        if not comboio:
            comboio = tanque_movel
        
        transferencias = Transferencia.objects.filter(movel__in=comboio, fixo__in=tanque, data__range=[data_inicio, data_fim])
            


    return render(request, 'transferencias.html', {'transferencias': transferencias, 'tanque_fixo':tanque_fixo, 'tanque_movel':tanque_movel})
 else: return HttpResponse('<h1>Acesso Negado</h1>')


def obras(request):

 if request.user.status == "c":
    obras = Obras.objects.all()
    saidas = []
    entradas = []
    
    for obra in obras:
        metodo_saidas = Abastecimento.objects.filter(obra=obra).aggregate(Sum('litros'))['litros__sum']
        metodo_entradas = Entrada.objects.filter(obra=obra).aggregate(Sum('quantidade'))['quantidade__sum']
        if metodo_saidas == None:
            metodo_saidas = 0
        if metodo_entradas == None:
            metodo_entradas = 0

        obra.saldo = metodo_entradas - metodo_saidas
        obra.save()
        saidas.append(metodo_saidas)
        entradas.append(metodo_entradas)

    my_list = zip(obras, saidas, entradas)
    return render(request, 'obras.html', {'my_list': my_list})
 else: return HttpResponse("<h1>Acesso Negado</h1>")

def painel_obras(request):
 if request.user.status == "c":

    saidas = Abastecimento.objects.filter(status=False)

    return render(request, "painel_obras.html", {"saidas":saidas})
 else: return HttpResponse("<h1>Acesso Negado</h1>")


def obras_ano(request):
 if request.user.status == 'c':
    if request.method == "GET":
        year = datetime.today().year
    elif request.method == "POST":
        year = request.POST.get("ano")

    g_vunit = grafico_vunit(year)    

    yearb = int(year)-1
    lista =[]
    obras = Obras.objects.all()
    for obra in obras:
        l=[]
        l.append(obra.nome)
        for n in range(1,13):
            total_saidas = nonetest(Abastecimento.objects.filter(obra=obra).filter(data__year=year).filter(data__month=n).aggregate(Sum('litros'))['litros__sum'])
            total_obra = nonetest(Abastecimento.objects.filter(obra=obra).filter(data__year=year).aggregate(Sum('litros'))['litros__sum'])
            l.append(total_saidas)
        l.append(total_obra)
        lista.append(l)  


    saidas_meses = []
    entradas_meses = []
    entradas_meses_valor = []
    preco_unitario_meses =[]
    valor_saidas_meses =[]
    saldo_meses =[]
    saldo_acumulado_meses =[]
    for m in range(1,13):    
        saidas_mes = nonetest(Abastecimento.objects.filter(data__month=m).filter(data__year=year).aggregate(Sum('litros'))['litros__sum'])
        entradas_mes = nonetest(Entrada.objects.filter(data_entrega__month=m).filter(data_entrega__year=year).aggregate(Sum('quantidade'))['quantidade__sum'])
        entradas_mes_valor = nonetest(Entrada.objects.filter(data_entrega__month=m).filter(data_entrega__year=year).aggregate(Sum('preco_total'))['preco_total__sum'])
        preco_unitario_mes = nonetest(Entrada.objects.filter(data_entrega__month=m).filter(data_entrega__year=year).aggregate(Avg('preco_unitario'))['preco_unitario__avg'])
        
        saidas_meses.append(saidas_mes)
        entradas_meses.append(entradas_mes)
        entradas_meses_valor.append(entradas_mes_valor)
        preco_unitario_meses.append(preco_unitario_mes)
    
    for i in range(0, len(saidas_meses)):
        valor_saidas_meses.append(saidas_meses[i]*preco_unitario_meses[i])
        saldo_meses.append(entradas_meses[i]-saidas_meses[i])
    valor_saidas_ano = sum(valor_saidas_meses)
    saldo_ano =sum(saldo_meses)


    saidas_ano = nonetest(Abastecimento.objects.filter(data__year=year).aggregate(Sum('litros'))['litros__sum'])
    entradas_ano = nonetest(Entrada.objects.filter(data_entrega__year=year).aggregate(Sum('quantidade'))['quantidade__sum'])
    entradas_ano_valor = nonetest(Entrada.objects.filter(data_entrega__year=year).aggregate(Sum('preco_total'))['preco_total__sum'])
    preco_unitario_ano = nonetest(Entrada.objects.filter(data_entrega__year=year).aggregate(Avg('preco_unitario'))['preco_unitario__avg'])
    saldo_ano_anterior = Saldo.objects.get(ano=yearb) #coletar um específico objeto
    saldo_ano_anteriorv = Saldo.objects.filter(ano=yearb).aggregate(Sum('quantidade'))['quantidade__sum']
    saldo_acumulado_ano = float(entradas_ano) - float(saidas_ano) + float(saldo_ano_anteriorv)
   
    count = 0
    for i in saldo_meses:
        
        
        if count == 0:
            saldo_acumulado_meses.append(i + saldo_ano_anteriorv)

        else:
            saldo_acumulado_meses.append(i + saldo_acumulado_meses[count-1])
        count+=1
            
    return render(request, 'obras_ano.html', {'obras':obras, 
                                              'lista':lista, 
                                              'saidas_meses':saidas_meses, 
                                              'saidas_ano':saidas_ano, 
                                              'entradas_meses':entradas_meses, 
                                              'entradas_ano':entradas_ano,
                                              'entradas_meses_valor':entradas_meses_valor,
                                              'entradas_ano_valor':entradas_ano_valor,
                                              'preco_unitario_meses':preco_unitario_meses,
                                              'preco_unitario_ano':preco_unitario_ano,
                                              'valor_saidas_meses':valor_saidas_meses,
                                              'valor_saidas_ano':valor_saidas_ano,
                                              'saldo_meses':saldo_meses,
                                              'saldo_ano':saldo_ano,
                                              'yearb':yearb,
                                              'saldo_ano_anterior':saldo_ano_anterior,
                                              'saldo_acumulado_meses':saldo_acumulado_meses,
                                              'saldo_acumulado_ano':saldo_acumulado_ano,
                                              'grafico_vunit':g_vunit})

 else: return HttpResponse("<h1>Acesso Negado</h1>")
def importexcel(request):
    if request.method == 'POST':
        
        if 'excel' in request.FILES:
            excelfile = request.FILES['excel']
            workbook = openpyxl.load_workbook(excelfile)
            saidas = workbook['SAIDAS']
            entradas = workbook['ENTRADAS']
            transferencias = workbook['TRANSFERENCIAS']
            user_id = request.user

            #numeros:

            
            for i in entradas.iter_rows(min_row=3,values_only=True):
                
                ent = i[:10]
                print(ent)
                if ent[0] == None:
                        break
                else:
                
                        tanque = Tanque.objects.filter(prefixo=ent[0])[0]
                        obra = Obras.objects.filter(nome=ent[8])[0]
                        num_entrada = Entrada.objects.aggregate(Max('numero'))
                        num_entrada = num_entrada['numero__max']+1

                        entrada = Entrada(tanque=tanque,
                                            nota_fiscal=ent[4],
                                            fornecedor=ent[3],
                                            data_nf=ent[5],
                                            data_entrega=ent[1],
                                            obra=obra,
                                            quantidade=ent[2],
                                            preco_unitario=ent[6],
                                            preco_total=ent[7],
                                            colaborador=user_id,
                                            descricao=ent[9],
                                            numero = num_entrada)
                                            
                        entrada.save() 

            for i in transferencias.iter_rows(min_row=3, values_only=True):
                    
                    
                    tr = i[:7]
                    if tr[0] == None:
                        break
                    else:
                        tanque_fixo = Tanque.objects.get(prefixo=tr[0])
                        tanque_movel= Tanque.objects.get(prefixo=tr[1])


                        transferencia = Transferencia(fixo=tanque_fixo,
                                        movel=tanque_movel,
                                        contador_inicio=tr[2],
                                        contador_fim=tr[3],
                                        litros=tr[3]-tr[2],
                                        contador_comboio=tr[6],
                                        colaborador=user_id,
                                        data=tr[5])
                        transferencia.save()

            for i in saidas.iter_rows(min_row=3,values_only=True):
                    sds = i[:10]
                    print(sds)

                    if sds[0] == None:
                        break
                    else:
                        tanque_saida = Tanque.objects.get(prefixo=sds[0])
                        obra_saida = Obras.objects.get(nome=sds[2])
                        try:
                            equipamento_saida = Equipamentos.objects.get(prefixo=sds[3])
                            if sds[8] == None: lubrificacao = False
                            else: lubrificacao = True

                            if type(sds[7]) != int and type(sds[7]) != float:
                                horimetro = 0
                            else: horimetro =sds[7]

                        except:
                            print(sds[3], "erro")



                        num_saida = Abastecimento.objects.aggregate(Max('numero')) 
                        num_saida = (num_saida["numero__max"]+1)
                        print(num_saida)

                        saida = Abastecimento(litros=sds[5]-sds[4],
                                            contador_inicio=sds[4],
                                            contador_fim=sds[5],
                                            horimetro=horimetro,
                                            data=sds[1],
                                            tanque=tanque_saida,
                                            obra=obra_saida,
                                            equipamento=equipamento_saida,
                                            lubrificacao=lubrificacao,
                                            operador=sds[9],
                                            colaborador=user_id,
                                            status=True,
                                            observacao="",
                                            numero = num_saida
                                            )
                        
                        try:
                            saida.save()  
                            messages.add_message(request, constants.SUCCESS, "Arquivo importado com sucesso!" )  
                        except:
                            messages.add_message(request, constants.ERROR, "Erro na saída!" ) 

    else:
        pass
    return redirect('home/')




